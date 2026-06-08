"""
点单系统 v1 - Flask 单文件后端
对应设计文档：DESIGN.md
"""
import json
import os
import re
import secrets
import string
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import xlrd
from flask import Flask, abort, jsonify, render_template, request, send_file
from werkzeug.middleware.proxy_fix import ProxyFix

BASE_DIR = Path(__file__).parent
MENU_PATH = BASE_DIR / "data" / "menu.json"

app = Flask(__name__)


class ScriptNameMiddleware:
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        script_name = environ.get("HTTP_X_SCRIPT_NAME", "")
        if script_name:
            environ["SCRIPT_NAME"] = script_name
        return self.app(environ, start_response)


app.wsgi_app = ScriptNameMiddleware(app.wsgi_app)

# ---------- 数据加载 ----------
with MENU_PATH.open("r", encoding="utf-8") as f:
    MENU = json.load(f)

# 进程内存储，重启即清空
SESSIONS: dict[str, dict] = {}
ORDERS: dict[str, dict] = {}  # order_id -> order dict

# 会话码字符表（去掉易混淆的 0/O/1/I/L）
SESSION_CODE_ALPHABET = "".join(c for c in (string.ascii_uppercase + string.digits) if c not in "0O1IL")


# ---------- 工具函数 ----------
def gen_session_code() -> str:
    for _ in range(5):
        code = "".join(secrets.choice(SESSION_CODE_ALPHABET) for _ in range(6))
        if code not in SESSIONS:
            return code
    raise RuntimeError("无法生成唯一会话码")


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def get_session_or_404(session_id: str) -> dict:
    s = SESSIONS.get(session_id)
    if not s:
        abort(404, description="session not found")
    return s


def calc_unit_price(dish: dict, sauce_ids: list[str], addon_ids: list[str]) -> float:
    sauces = {s["id"]: s for s in MENU["sauces"]}
    addons = {a["id"]: a for a in MENU["addons"]}
    total = dish["price"]
    for sid in sauce_ids:
        total += sauces.get(sid, {}).get("price", 0)
    for aid in addon_ids:
        total += addons.get(aid, {}).get("price", 0)
    return round(total, 2)


def get_order_or_404(order_id: str) -> dict:
    o = ORDERS.get(order_id)
    if not o:
        abort(404, description="order not found")
    return o


# ---------- 页面路由 ----------
@app.get("/")
def page_index():
    return render_template("index.html")


@app.get("/menu/<session_id>")
def page_menu(session_id: str):
    get_session_or_404(session_id)
    return render_template("menu.html", session_id=session_id)


@app.get("/summary/<session_id>")
def page_summary(session_id: str):
    get_session_or_404(session_id)
    return render_template("summary.html", session_id=session_id)


@app.get("/admin/<session_id>")
def page_admin(session_id: str):
    get_session_or_404(session_id)
    return render_template("admin.html", session_id=session_id)


@app.get("/admin/sessions")
def page_sessions():
    return render_template("sessions.html")


# ---------- API ----------
@app.get("/api/menu")
def api_menu():
    return jsonify(MENU)


@app.post("/api/session")
def api_create_session():
    body = request.get_json(silent=True) or {}
    title = (body.get("title") or "").strip()
    sid = gen_session_code()
    SESSIONS[sid] = {
        "id": sid,
        "title": title or f"会话 {sid}",
        "created_at": now_iso(),
        "status": "active",
        "user_names": [],
        "order_ids": [],
    }
    return jsonify(SESSIONS[sid])


@app.get("/api/session/<session_id>")
def api_get_session(session_id: str):
    s = get_session_or_404(session_id)
    orders = [ORDERS[oid] for oid in s["order_ids"]]
    return jsonify({**s, "orders": orders})


@app.post("/api/session/<session_id>/order")
def api_add_order(session_id: str):
    s = get_session_or_404(session_id)
    if s["status"] != "active":
        return jsonify({"error": "session closed"}), 400
    body = request.get_json(silent=True) or {}
    user_name = (body.get("user_name") or "").strip()
    dept = (body.get("dept") or "").strip()
    dish_id = body.get("dish_id")
    sauce_ids = body.get("sauce_ids") or []
    addon_ids = body.get("addon_ids") or []
    base = (body.get("base") or "").strip()
    note = (body.get("note") or "").strip()
    if not user_name:
        return jsonify({"error": "user_name required"}), 400
    dish = next((d for d in MENU["dishes"] if d["id"] == dish_id), None)
    if not dish:
        return jsonify({"error": "dish not found"}), 400
    if dish.get("sauce_mode") == "single_required":
        if not isinstance(sauce_ids, list) or len(sauce_ids) != 1:
            return jsonify({"error": "sauce required (pick exactly 1)"}), 400
    base_options = dish.get("base_options")
    if base_options:
        if not base or base not in base_options:
            return jsonify({"error": "base required (pick 1 from base_options)"}), 400
    unit_price = calc_unit_price(dish, sauce_ids, addon_ids)
    # 分配用户序号：按首次出现的昵称顺序
    if user_name not in s["user_names"]:
        s["user_names"].append(user_name)
    user_no = s["user_names"].index(user_name) + 1
    # seq：取当前会话最大 seq + 1
    existing_seqs = [ORDERS[oid].get("seq", 0) for oid in s["order_ids"] if oid in ORDERS]
    next_seq = max(existing_seqs, default=0) + 1
    order = {
        "id": uuid.uuid4().hex[:12],
        "session_id": session_id,
        "user_name": user_name,
        "dept": dept,
        "user_no": user_no,
        "seq": next_seq,
        "dish_id": dish_id,
        "base": base,
        "sauce_ids": sauce_ids,
        "addon_ids": addon_ids,
        "note": note,
        "unit_price": unit_price,
        "created_at": now_iso(),
    }
    ORDERS[order["id"]] = order
    s["order_ids"].append(order["id"])
    return jsonify(order), 201


@app.delete("/api/order/<order_id>")
def api_delete_order(order_id: str):
    order = get_order_or_404(order_id)
    session = SESSIONS.get(order["session_id"])
    if session and order["id"] in session["order_ids"]:
        session["order_ids"].remove(order["id"])
    ORDERS.pop(order_id, None)
    return jsonify({"ok": True})


@app.get("/api/sessions")
def api_list_sessions():
    sessions = sorted(SESSIONS.values(), key=lambda s: s["created_at"], reverse=True)
    result = []
    for s in sessions:
        orders = [ORDERS[oid] for oid in s["order_ids"] if oid in ORDERS]
        result.append({**s, "order_count": len(orders)})
    return jsonify(result)


@app.delete("/api/session/<session_id>")
def api_delete_session(session_id: str):
    s = get_session_or_404(session_id)
    for oid in s.get("order_ids", []):
        ORDERS.pop(oid, None)
    SESSIONS.pop(session_id, None)
    return jsonify({"ok": True})


@app.delete("/api/sessions")
def api_clear_sessions():
    ORDERS.clear()
    SESSIONS.clear()
    return jsonify({"ok": True})


def _filter_sessions():
    """根据请求参数 ids 过滤会话，未传则返回全部"""
    ids = request.args.get("ids", "")
    if ids:
        id_list = [i.strip() for i in ids.split(",") if i.strip()]
        return {sid: s for sid, s in SESSIONS.items() if sid in id_list}
    return dict(SESSIONS)


@app.get("/api/export")
def api_export():
    sessions = _filter_sessions()
    if not sessions:
        return jsonify({"error": "没有可导出的会话"}), 400

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sauce_map = {s["id"]: s["name"] for s in MENU["sauces"]}

    # 汇总会话排最前
    sorted_sessions = sorted(sessions.items(), key=lambda x: ("汇总" not in x[1]["title"], x[1]["title"]))
    for sid, session in sorted_sessions:
        orders = [ORDERS[oid] for oid in session["order_ids"] if oid in ORDERS]
        if not orders:
            continue

        # 聚合：与 admin 页面 renderItems 逻辑一致
        map = {}

        def bump(key, name, tag, no):
            if key not in map:
                map[key] = {"name": name, "tag": tag, "count": 0, "user_nos": set()}
            map[key]["count"] += 1
            map[key]["user_nos"].add(no)

        for o in orders:
            d = next((x for x in MENU["dishes"] if x["id"] == o["dish_id"]), {})
            display = d.get("name", "?") + (f"（{o['base']}）" if o.get("base") else "")
            no = o.get("seq", o["user_no"])
            bump("dish:" + o["dish_id"] + ":" + (o.get("base") or ""), display, "主食", no)
            for sid_s in o.get("sauce_ids", []):
                bump("sauce:" + sid_s, sauce_map.get(sid_s, "?"), "酱料", no)

        dish_rows = sorted(
            [r for r in map.values() if r["tag"] == "主食"],
            key=lambda r: (-r["count"], r["name"]),
        )
        sauce_rows = sorted(
            [r for r in map.values() if r["tag"] == "酱料"],
            key=lambda r: (-r["count"], r["name"]),
        )

        # Excel sheet 名最多 31 字，不能含 / \ [ ] : *
        sheet_name = re.sub(r'[/\\[\]:*?]', '-', session["title"])[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.append(["项目", "数量", "序号"])
        header_fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFD6E4F0")
        header_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="999999"),
            right=openpyxl.styles.Side(style="thin", color="999999"),
            top=openpyxl.styles.Side(style="thin", color="999999"),
            bottom=openpyxl.styles.Side(style="thin", color="999999"),
        )
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = header_border
        for r in dish_rows:
            nos = ", ".join(str(n) for n in sorted(r["user_nos"]))
            ws.append([r["name"], r["count"], nos])
        for r in sauce_rows:
            nos = ", ".join(str(n) for n in sorted(r["user_nos"]))
            ws.append([r["name"], r["count"], nos])

        # 合计行
        total_count = sum(r["count"] for r in dish_rows) + sum(r["count"] for r in sauce_rows)
        ws.append(["合计", total_count, ""])

        # 动态列宽：中文字符占2宽度，英文/数字占1宽度
        def _char_width(c):
            return 2 if ord(c) > 127 else 1
        def _val_width(val):
            return sum(_char_width(c) for c in str(val or ""))
        max_a = max(_val_width(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1))
        max_b = max(_val_width(ws.cell(row=r, column=2).value) for r in range(1, ws.max_row + 1))
        max_c = max(_val_width(ws.cell(row=r, column=3).value) for r in range(1, ws.max_row + 1))
        ws.column_dimensions["A"].width = max_a + 4
        ws.column_dimensions["B"].width = max_b + 7
        ws.column_dimensions["C"].width = 40

        # 行高 + 居中 + 边框 + 隔行灰色
        center_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        wrap_align = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = openpyxl.styles.Side(style="thin", color="999999")
        all_border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        gray_fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFF2F2F2")
        col_c_width = ws.column_dimensions["C"].width
        chars_per_line = int(col_c_width)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
            row_idx = row[0].row
            c_val = str(ws.cell(row=row_idx, column=3).value or "")
            char_len = _val_width(c_val)
            lines = max(1, -(-char_len // chars_per_line))
            ws.row_dimensions[row_idx].height = max(22, lines * 18)
            for cell in row:
                cell.alignment = wrap_align if cell.column == 3 else center_align
                cell.border = all_border
                if row_idx > 1 and row_idx % 2 == 0:
                    cell.fill = gray_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="统计汇总.xlsx",
    )


@app.get("/api/export/detail")
def api_export_detail():
    """导出明细表，格式与导入的原始 Excel 一致"""
    sessions = _filter_sessions()
    if not sessions:
        return jsonify({"error": "没有可导出的会话"}), 400

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sauce_map = {s["id"]: s["name"] for s in MENU["sauces"]}
    dish_map = {d["id"]: d for d in MENU["dishes"]}

    for sid, session in sessions.items():
        orders = sorted(
            [ORDERS[oid] for oid in session["order_ids"] if oid in ORDERS],
            key=lambda o: o.get("seq", o.get("user_no", 0)),
        )
        if not orders:
            continue

        # 从会话标题提取园区和日期
        area = session["title"].split(" - ")[-1] if " - " in session["title"] else ""
        title_prefix = session["title"].split(" - ")[0] if " - " in session["title"] else session["title"]

        # 标题格式：三德科技团餐-2026/06/03
        dm = re.search(r'(\d+)月(\d+)日', title_prefix)
        if dm:
            now = datetime.now()
            date_str = f"{now.year}/{int(dm.group(1)):02d}/{int(dm.group(2)):02d}"
        else:
            date_str = title_prefix
        excel_title = f"三德科技团餐-{date_str}"

        sheet_name = area or excel_title
        ws = wb.create_sheet(title=sheet_name)

        # Row 1: 标题（合并单元格）
        ws.append([excel_title, None, None, None, None, None, None])
        ws.merge_cells("A1:G1")
        ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        ws["A1"].border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="999999"),
            right=openpyxl.styles.Side(style="thin", color="999999"),
            top=openpyxl.styles.Side(style="thin", color="999999"),
            bottom=openpyxl.styles.Side(style="thin", color="999999"),
        )
        # Row 2: 表头（蓝色背景 + 四边框线）
        ws.append(["序号", "部门", "姓名", "餐品名称", "酱汁", "金额", "就餐园区"])
        header_fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFD6E4F0")
        header_border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="999999"),
            right=openpyxl.styles.Side(style="thin", color="999999"),
            top=openpyxl.styles.Side(style="thin", color="999999"),
            bottom=openpyxl.styles.Side(style="thin", color="999999"),
        )
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = header_border

        # 数据行
        grand_total = 0
        for o in orders:
            dish = dish_map.get(o["dish_id"], {})
            dish_name = dish.get("name", "?")
            if o.get("base"):
                dish_name += f"（{o['base']}）"
            sauce_name = sauce_map.get(o["sauce_ids"][0], "") if o.get("sauce_ids") else ""
            ws.append([
                o.get("seq", o["user_no"]),
                o.get("dept", ""),
                o["user_name"],
                dish_name,
                sauce_name,
                o["unit_price"],
                o.get("area", area),
            ])
            grand_total += o["unit_price"]

        # 汇总行 + 送餐门店
        ws.append(["送餐门店：适绿轻食中电店", None, None, None, "合计/元", grand_total, None])
        ws.append(["送餐人：", None, None, "签收人：", "时间", None, None])

        # 合并单元格
        summary_row = ws.max_row - 1
        last_row = ws.max_row
        for r in (summary_row, last_row):
            ws.merge_cells(f"A{r}:C{r}")
            ws.merge_cells(f"F{r}:G{r}")
            ws.row_dimensions[r].height = 28

        # 行高 + 对齐 + 全表边框
        center_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        left_align = openpyxl.styles.Alignment(horizontal="left", vertical="center")
        left_indent = openpyxl.styles.Alignment(horizontal="left", vertical="center", indent=2)
        thin = openpyxl.styles.Side(style="thin", color="999999")
        all_border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
        left_only = openpyxl.styles.Border(left=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if row[0].row not in (summary_row, last_row):
                ws.row_dimensions[row[0].row].height = 22
            for cell in row:
                cell.border = all_border
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                r, c = cell.row, cell.column
                if r == summary_row:
                    cell.alignment = left_indent if c == 1 else center_align
                elif r == last_row:
                    cell.alignment = left_indent if c in (1, 4) else (left_align if c in (6, 7) else center_align)
                else:
                    cell.alignment = center_align
        # 确保最左侧边框完整
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=1).border = all_border

        # 列宽自适应（跳过合并单元格）
        for col_idx in range(1, 8):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                val = str(cell.value or "")
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        # 序号列固定窄宽度
        ws.column_dimensions["A"].width = 8

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="明细统计.xlsx",
    )


# ---------- Excel 导入 ----------
def _char_overlap(a: str, b: str) -> float:
    """计算两个字符串的字符重叠率"""
    if not a or not b:
        return 0
    overlap = sum(1 for c in a if c in b)
    return overlap / max(len(a), len(b))


def _match_dish(name: str) -> dict | None:
    """模糊匹配菜品名称，返回 menu 中的 dish dict"""
    # 1. 精确匹配
    for d in MENU["dishes"]:
        if d["name"] == name:
            return d
    # 2. 子串匹配
    for d in MENU["dishes"]:
        if d["name"] in name or name in d["name"]:
            return d
    # 3. 去掉常见品类后缀再匹配
    suffixes = ["碗", "面", "饭", "沙拉", "意面"]
    for d in MENU["dishes"]:
        for suf in suffixes:
            if name.endswith(suf) and d["name"] == name[: -len(suf)]:
                return d
            if d["name"].endswith(suf) and name == d["name"][: -len(suf)]:
                return d
    # 4. 字符重叠兜底
    best, best_score = None, 0
    for d in MENU["dishes"]:
        score = _char_overlap(name, d["name"])
        if score > best_score:
            best_score = score
            best = d
    if best_score >= 0.6:
        return best
    return None


def _match_sauce(name: str) -> str | None:
    """模糊匹配酱汁名称，返回 sauce id"""
    if not name:
        return None
    name = name.strip()
    # 1. 精确匹配
    for s in MENU["sauces"]:
        if s["name"] == name:
            return s["id"]
    # 2. 子串匹配
    for s in MENU["sauces"]:
        if name in s["name"] or s["name"] in name:
            return s["id"]
    # 3. 去掉通用后缀，用关键词子串匹配
    suffixes = ["沙拉汁", "芝麻酱", "沙拉酱", "甜辣酱", "辣酱", "酱", "汁"]
    core = name
    for suf in suffixes:
        if core.endswith(suf) and len(core) > len(suf):
            core = core[: -len(suf)]
            break
    # 先试 2 字子串
    if len(core) >= 2:
        for i in range(len(core) - 1):
            seg = core[i : i + 2]
            for s in MENU["sauces"]:
                if seg in s["name"]:
                    return s["id"]
    # 再试逐字匹配（core 本身就短或者 2 字子串都没命中时）
    for c in core:
        for s in MENU["sauces"]:
            if c in s["name"]:
                return s["id"]
    # 4. 字符重叠兜底
    best_id, best_score = None, 0
    for s in MENU["sauces"]:
        score = _char_overlap(name, s["name"])
        if score > best_score:
            best_score = score
            best_id = s["id"]
    if best_score >= 0.6:
        return best_id
    return None


def _parse_dish_base(cell: str) -> tuple[str, str]:
    """从 '鸡胸肉牛肉双拼碗（荞麦面）' 解析出 (dish_name, base)"""
    m = re.match(r'^(.+?)[\(（](.+?)[\)）]$', cell.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return cell.strip(), ""


def _create_session(title: str) -> tuple[str, list]:
    """创建一个会话，返回 (session_id, warnings_list)"""
    sid = gen_session_code()
    SESSIONS[sid] = {
        "id": sid,
        "title": title,
        "created_at": now_iso(),
        "status": "active",
        "user_names": [],
        "order_ids": [],
        "import_warnings": [],
    }
    return sid, SESSIONS[sid]["import_warnings"]


def _import_order(sid: str, warnings: list, user_name: str, dish_cell: str, sauce_cell: str, base: str, seq: int = 0, dept: str = ""):
    """解析并导入一条订单"""
    if not user_name or not dish_cell:
        return

    dish = _match_dish(dish_cell)
    if not dish:
        warnings.append({"type": "dish", "original": dish_cell, "matched": None, "user": user_name, "no": seq})
        return
    elif dish["name"] != dish_cell:
        warnings.append({"type": "dish", "original": dish_cell, "matched": dish["name"], "user": user_name, "no": seq})

    sauce_id = _match_sauce(sauce_cell)
    if sauce_cell:
        if sauce_id:
            sauce_name = next(s["name"] for s in MENU["sauces"] if s["id"] == sauce_id)
            if sauce_name != sauce_cell:
                warnings.append({"type": "sauce", "original": sauce_cell, "matched": sauce_name, "user": user_name, "no": seq})
        else:
            warnings.append({"type": "sauce", "original": sauce_cell, "matched": None, "user": user_name, "no": seq})
    sauce_ids = [sauce_id] if sauce_id else []

    session = SESSIONS[sid]
    if user_name not in session["user_names"]:
        session["user_names"].append(user_name)
    user_no = session["user_names"].index(user_name) + 1

    unit_price = calc_unit_price(dish, sauce_ids, [])
    order = {
        "id": uuid.uuid4().hex[:12],
        "session_id": sid,
        "user_name": user_name,
        "dept": dept,
        "user_no": user_no,
        "seq": seq,
        "dish_id": dish["id"],
        "base": base,
        "sauce_ids": sauce_ids,
        "addon_ids": [],
        "note": "",
        "unit_price": unit_price,
        "created_at": now_iso(),
    }
    ORDERS[order["id"]] = order
    session["order_ids"].append(order["id"])


def _detect_format(headers: tuple) -> str:
    """根据表头行判断格式：'v1' 每sheet一会话, 'v2' 按园区分组"""
    # v2 格式：E列=主食, G列=就餐园区
    if len(headers) >= 7 and headers[4] and "主食" in str(headers[4]):
        return "v2"
    return "v1"


def _parse_seq(val) -> int | None:
    """将序号值转为 int，支持 int/float/str，无效返回 None"""
    if isinstance(val, (int, float)):
        return int(val) if val else None
    s = str(val).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, OverflowError):
        return None


def _import_v1_sheet(sheet_name: str, rows: list) -> dict:
    """格式1：每个 sheet 一个会话，base 写在菜品名括号里"""
    title = str(rows[0][0] or "").strip()
    title = f"{title} - {sheet_name}" if title else sheet_name

    sid, warnings = _create_session(title)

    for row in rows[2:]:
        seq = _parse_seq(row[0])
        if not seq:
            break
        user_name = str(row[2] or "").strip()
        dish_cell = str(row[3] or "").strip()
        sauce_cell = str(row[4] or "").strip()
        if not user_name or not dish_cell:
            continue
        dept = str(row[1] or "").strip()
        dish_name, base = _parse_dish_base(dish_cell)
        _import_order(sid, warnings, user_name, dish_name, sauce_cell, base, seq, dept)

    return {"id": sid, "title": title, "warnings": len(warnings)}


def _import_v2_sheet(sheet_name: str, rows: list) -> list[dict]:
    """格式2：单 sheet，按就餐园区列分组生成多个会话"""
    date_title = str(rows[0][0] or "").strip()
    if not date_title:
        date_title = sheet_name

    # 按园区分组
    groups: dict[str, list] = {}
    for row in rows[2:]:
        seq = _parse_seq(row[0])
        if not seq:
            break
        area = str(row[6] or "").strip() if len(row) > 6 else ""
        if not area:
            area = "未分组"
        groups.setdefault(area, []).append(row)

    results = []
    for area, area_rows in groups.items():
        title = f"{date_title} - {area}"
        sid, warnings = _create_session(title)

        for row in area_rows:
            user_name = str(row[2] or "").strip()
            dish_cell = str(row[3] or "").strip()
            base = str(row[4] or "").strip() if len(row) > 4 else ""
            sauce_cell = str(row[5] or "").strip() if len(row) > 5 else ""
            dept = str(row[1] or "").strip()
            _import_order(sid, warnings, user_name, dish_cell, sauce_cell, base, _parse_seq(row[0]) or 0, dept)

        results.append({"id": sid, "title": title, "warnings": len(warnings)})

    return results


def _sheet_date(sheet_name: str) -> tuple[int, int]:
    """从 sheet 名解析日期，如 '6.1-6.5' → (6, 1)，用于排序取最新"""
    m = re.search(r'(\d+)\.(\d+)', sheet_name)
    if m:
        return int(m.group(1)), int(m.group(2))
    return 0, 0


def _sheet_menu_score(rows: list) -> float:
    """统计 sheet 中菜品名与当前菜单的匹配率"""
    dish_names = {d["name"] for d in MENU["dishes"]}
    total, matched = 0, 0
    for row in rows[2:]:
        dish_cell = str(row[3] or "").strip() if len(row) > 3 else ""
        if not dish_cell:
            continue
        total += 1
        if dish_cell in dish_names or any(d in dish_cell or dish_cell in d for d in dish_names):
            matched += 1
    return matched / total if total else 0


@app.post("/api/import")
def api_import_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "请上传 Excel 文件"}), 400

    filename = (file.filename or "").lower()
    if not (filename.endswith(".xls") or filename.endswith(".xlsx")):
        return jsonify({"error": "仅支持 .xls 或 .xlsx 格式"}), 400

    try:
        if filename.endswith(".xls"):
            wb = xlrd.open_workbook(file_contents=file.read())
            sheets = []
            for name in wb.sheet_names():
                ws = wb.sheet_by_name(name)
                rows = []
                for i in range(ws.nrows):
                    rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
                sheets.append((name, rows))
        else:
            wb = openpyxl.load_workbook(file, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets.append((name, list(ws.iter_rows(min_row=1, values_only=True))))
    except Exception:
        return jsonify({"error": "无法解析 Excel 文件"}), 400

    # 筛选候选 sheet：跳过菜单 sheet 和数据不足的 sheet
    candidates = []
    for sheet_name, rows in sheets:
        if len(rows) < 3:
            continue
        if "菜单" in sheet_name:
            continue
        headers = rows[1]
        if not (len(headers) >= 5 and "序号" in str(headers[0])):
            continue
        candidates.append((sheet_name, rows))

    if not candidates:
        return jsonify({"error": "未找到有效的订单数据 sheet"}), 400

    # 选最佳 sheet：菜单匹配率最高的，相同时取最新日期
    def _score(item):
        name, rows = item
        return (_sheet_menu_score(rows), *_sheet_date(name))

    best_name, best_rows = max(candidates, key=_score)

    headers = best_rows[1]
    fmt = _detect_format(headers)
    if fmt == "v2":
        created = _import_v2_sheet(best_name, best_rows)
    else:
        created = [_import_v1_sheet(best_name, best_rows)]

    if not created:
        return jsonify({"error": "未解析到有效数据"}), 400

    # 创建汇总会话：合并本次导入的所有订单
    all_orders = []
    for item in created:
        s = SESSIONS[item["id"]]
        for oid in s["order_ids"]:
            all_orders.append(ORDERS[oid])

    if all_orders:
        # 取已创建会话标题的前缀，拼接"汇总"
        first_title = created[0]["title"]
        prefix = first_title.rsplit(" - ", 1)[0] if " - " in first_title else first_title
        # 匹配麓谷或长兴时加公司名
        all_titles = " ".join(item["title"] for item in created)
        company = "三德科技 " if ("麓谷" in all_titles or "长兴" in all_titles) else ""
        summary_title = prefix + " - " + company + "汇总"
        sid, summary_warnings = _create_session(summary_title)
        # 合并各会话的导入匹配结果
        for item in created:
            src = SESSIONS.get(item["id"])
            if src:
                summary_warnings.extend(src.get("import_warnings", []))
        for o in all_orders:
            # 从原会话标题提取园区，写入订单
            orig_session = SESSIONS.get(o["session_id"])
            orig_title = orig_session["title"] if orig_session else ""
            orig_area = orig_title.split(" - ")[-1] if " - " in orig_title else ""
            new_order = {**o, "id": uuid.uuid4().hex[:12], "session_id": sid, "area": orig_area}
            ORDERS[new_order["id"]] = new_order
            SESSIONS[sid]["order_ids"].append(new_order["id"])
            if new_order["user_name"] not in SESSIONS[sid]["user_names"]:
                SESSIONS[sid]["user_names"].append(new_order["user_name"])
        created.append({"id": sid, "title": summary_title, "warnings": len(summary_warnings)})

    return jsonify({"sessions": created}), 201


@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": str(e.description)}), 404


def _import_file(filepath: str):
    """从文件路径导入 Excel，复用与 /api/import 相同的逻辑"""
    filename = os.path.basename(filepath).lower()
    with open(filepath, "rb") as f:
        if filename.endswith(".xls"):
            wb = xlrd.open_workbook(file_contents=f.read())
            sheets = []
            for name in wb.sheet_names():
                ws = wb.sheet_by_name(name)
                rows = []
                for i in range(ws.nrows):
                    rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
                sheets.append((name, rows))
        else:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                sheets.append((name, list(ws.iter_rows(min_row=1, values_only=True))))

    candidates = []
    for sheet_name, rows in sheets:
        if len(rows) < 3:
            continue
        if "菜单" in sheet_name:
            continue
        headers = rows[1]
        if not (len(headers) >= 5 and "序号" in str(headers[0])):
            continue
        candidates.append((sheet_name, rows))

    if not candidates:
        return

    def _score(item):
        name, rows = item
        return (_sheet_menu_score(rows), *_sheet_date(name))

    best_name, best_rows = max(candidates, key=_score)
    headers = best_rows[1]
    fmt = _detect_format(headers)
    if fmt == "v2":
        created = _import_v2_sheet(best_name, best_rows)
    else:
        created = [_import_v1_sheet(best_name, best_rows)]

    if not created:
        return

    all_orders = []
    for item in created:
        s = SESSIONS[item["id"]]
        for oid in s["order_ids"]:
            all_orders.append(ORDERS[oid])

    if all_orders:
        first_title = created[0]["title"]
        prefix = first_title.rsplit(" - ", 1)[0] if " - " in first_title else first_title
        all_titles = " ".join(item["title"] for item in created)
        company = "三德科技 " if ("麓谷" in all_titles or "长兴" in all_titles) else ""
        summary_title = prefix + " - " + company + "汇总"
        sid, summary_warnings = _create_session(summary_title)
        # 合并各会话的导入匹配结果
        for item in created:
            src = SESSIONS.get(item["id"])
            if src:
                summary_warnings.extend(src.get("import_warnings", []))
        for o in all_orders:
            orig_session = SESSIONS.get(o["session_id"])
            orig_title = orig_session["title"] if orig_session else ""
            orig_area = orig_title.split(" - ")[-1] if " - " in orig_title else ""
            new_order = {**o, "id": uuid.uuid4().hex[:12], "session_id": sid, "area": orig_area}
            ORDERS[new_order["id"]] = new_order
            SESSIONS[sid]["order_ids"].append(new_order["id"])
            if new_order["user_name"] not in SESSIONS[sid]["user_names"]:
                SESSIONS[sid]["user_names"].append(new_order["user_name"])

    print(f"  [auto-import] {os.path.basename(filepath)} -> {len(created)} 个会话")


def _auto_import_testfiles():
    """启动时自动导入 testfile 目录下的 Excel 文件"""
    testfile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "testfile")
    if not os.path.isdir(testfile_dir):
        return
    files = sorted(f for f in os.listdir(testfile_dir) if f.endswith((".xls", ".xlsx")))
    if not files:
        return
    print(f"=== 自动导入 {len(files)} 个测试文件 ===")
    for f in files:
        try:
            _import_file(os.path.join(testfile_dir, f))
        except Exception as e:
            print(f"  [auto-import] {f} 导入失败: {e}")


_auto_import_testfiles()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=35001, debug=True)
